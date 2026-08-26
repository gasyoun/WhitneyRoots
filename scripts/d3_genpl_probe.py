#!/usr/bin/env python3
"""D3 gen.pl accent probe — long-ī/ū gen.pl accent placement in the accented RV.

Data: Zurich morphologically glossed Rigveda (Casaretto et al. 2025), via the
VedaWebProject/vedaweb-data GitHub mirror (rigveda/versions/zurich.xlsx),
CC BY 4.0. Same resource as VedaWeb API corpus 66695e4a14f6d337f7788740.

Question (SL CONTRADICTIONS §1 / GAPS §1; WhitneyRoots D3): in the gen.pl of
derivative long-ī/ū feminine stems, is the accent thrown forward onto the
ending (-īnā́m, Whitney §319a, bahvī́-type adjectives) or kept on the stem
vowel (-ī́nām, Whitney §320/§356 paradigms, nadī́-type nouns)?

Two sections:
  A. lemma itself ends in long ī/ū (independent noun stems + a few roots)
  B. feminine tokens whose lemma does NOT end in long ī/ū — catches the
     devī́-declension feminines lemmatized under the masculine adjective /
     participle stem (bahú- f. bahvī́, aruṇá- f. aruṇī́, root participles),
     plus i/u-stem feminine nouns as a control.

Usage: python d3_genpl_probe.py [zurich.xlsx]
Output: report to stdout + d3_genpl_hits.json (all matched tokens).
"""
import sys, json, unicodedata, re
from collections import defaultdict
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

XLSX = sys.argv[1] if len(sys.argv) > 1 else 'zurich.xlsx'

ACUTE = '́'   # udātta acute in IAST
GRAVE = '̀'   # dependent svarita in some corpora; ignored for position

def nfc(s):
    return unicodedata.normalize('NFC', s or '')

def strip_accents(s):
    # NFC keeps ī/ū/ā precomposed; the acute stays combining (no precomposed
    # ī-acute exists) — removing combining acute/grave leaves a clean skeleton.
    return nfc(''.join(c for c in unicodedata.normalize('NFD', s or '')
                       if c not in (ACUTE, GRAVE)))

# ending shapes: -īnām / -ūnām, retroflexed -īṇām, sandhi -nām̐ variants
SKEL = re.compile('[īū][nṇ]ām?$')
CLASSIFY = re.compile('([īū])(́?)[nṇ](ā)(́?)ḿ?$')

def classify(form):
    d = nfc(form).replace('ṃ', 'm')   # anusvāra → m
    m = CLASSIFY.search(d)
    if not m:
        return 'unparsed'
    stem_acc, end_acc = bool(m.group(2)), bool(m.group(4))
    if stem_acc and not end_acc:
        return 'stem_final'        # -ī́nām (Whitney §320/§356)
    if end_acc and not stem_acc:
        return 'ending'            # -īnā́m (Whitney §319a)
    if not stem_acc and not end_acc:
        return 'accent_elsewhere'  # barytone: accent earlier in the word
    return 'both'

wb = load_workbook(XLSX, read_only=True)
ws = wb[wb.sheetnames[0]]
rows = ws.iter_rows(values_only=True)
header = next(rows)
idx = {h: i for i, h in enumerate(header)}
col = lambda r, name: r[idx[name]]

hits = []
gp_total = 0
for r in rows:
    if col(r, 'belege::numerus bestof') != 'Pl.' or col(r, 'belege::kasus bestof') != 'Gen.':
        continue
    gp_total += 1
    form = col(r, 'belege::form') or ''
    if not SKEL.search(strip_accents(form).replace('ṃ', 'm')):
        continue
    lemma = col(r, 'lemmata klassisch::lemma') or ''
    sk_lemma = strip_accents(lemma).rstrip('-')
    section = 'A' if sk_lemma and sk_lemma[-1] in 'īū' else 'B'
    hits.append({
        'section': section,
        'loc': col(r, 'belege::stelleMMSSSRR'), 'pada': col(r, 'belege::pada'),
        'form': form, 'lemma': lemma,
        'gender': col(r, 'belege::genus bestof'),
        'lemmatyp': col(r, 'lemmata klassisch::lemmatyp'),
        'meaning': (col(r, 'lemmata klassisch::bedeutung') or '')[:60],
        'class': classify(form),
    })

print(f"gen.pl tokens total in corpus: {gp_total}")
print(f"gen.pl tokens in long-ī/ū + nām shape: {len(hits)}")

def report(section, title, only_fem=False):
    sel = [h for h in hits if h['section'] == section
           and (not only_fem or (h['gender'] or '').lower().startswith('f'))]
    print(f"\n=== Section {section}: {title} ({len(sel)} tokens) ===")
    by_lemma = defaultdict(list)
    for h in sel:
        by_lemma[h['lemma']].append(h)
    for lemma in sorted(by_lemma):
        hs = by_lemma[lemma]
        cf = defaultdict(int)
        for h in hs:
            cf[(h['form'], h['class'])] += 1
        detail = '  '.join(f"{f} ×{n} [{c}]" for (f, c), n in sorted(cf.items()))
        print(f"{lemma}  ({hs[0]['gender']})  n={len(hs)}:  {detail}")
        print(f"    locs: {[h['loc'] for h in hs][:12]}")

report('A', 'lemma ends in long ī/ū (nadī́-type nouns, roots)')
report('B', 'FEMININE under non-ī/ū lemma (bahvī́-type adjectives/participles + i/u-stem noun control)',
       only_fem=True)

nonfem_b = defaultdict(int)
for h in hits:
    if h['section'] == 'B' and not (h['gender'] or '').lower().startswith('f'):
        nonfem_b[h['class']] += 1
print(f"\nSection B non-feminine control (masc/neut i/u-stems, Whitney §342 lengthening): {dict(nonfem_b)}")

with open('d3_genpl_hits.json', 'w', encoding='utf-8') as f:
    json.dump(hits, f, ensure_ascii=False, indent=1)
print("\nwrote d3_genpl_hits.json")
